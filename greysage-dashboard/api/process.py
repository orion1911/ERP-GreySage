import os
import json
import pandas as pd
import requests
from io import BytesIO
from datetime import datetime
import warnings
import openpyxl
import time
from http.server import BaseHTTPRequestHandler

warnings.filterwarnings('ignore', category=UserWarning)

# Configuration
ONEDRIVE_URL = os.getenv("ONEDRIVE_FILE_URL")
MAKER_SHEETS = ["GREYSAGE", "ARVIND", "MIDSEN", "HASAN", "RAMA", "HAKIM", "RAMU", "ANIL", "SINU"]

# Column indices we care about (found dynamically per sheet)
# Include DATE (column B) so we can filter rows by date
REQUIRED_COLS = {'CLIENT', 'WASHING', 'PCS', 'WASH ED', 'DATE', 'LOT NO.'}

# Stop reading after this many consecutive rows with no real data
MAX_EMPTY_STREAK = 20

# Global cache
_cache = {'data': None, 'timestamp': None}
CACHE_DURATION = 300


def download_excel_bytes():
    """Download Excel bytes with caching"""
    global _cache

    now = time.time()
    if _cache['data'] is not None and _cache['timestamp'] and now - _cache['timestamp'] < CACHE_DURATION:
        return _cache['data']

    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
    response = requests.get(ONEDRIVE_URL, timeout=20, headers=headers)
    response.raise_for_status()

    _cache['data'] = response.content
    _cache['timestamp'] = now
    return response.content


def process_all_sheets(excel_bytes):
    """
    Process all maker sheets using openpyxl read_only mode.
    
    KEY OPTIMIZATION: Some sheets (GREYSAGE, SINU) have 1M+ rows because
    column A is filled with the maker name all the way down, but only
    ~100 rows have actual multi-column data. We detect this and stop early,
    cutting read time from ~47s to <0.5s.
    """
    wb = openpyxl.load_workbook(BytesIO(excel_bytes), read_only=True, data_only=True)
    all_rows = []
    # keep records from November 2025 inclusive
    _date_cutoff = datetime(2025, 12, 9)

    try:
        for sheet_name in MAKER_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            sheet_rows = list(_read_sheet_fast(ws))

            if len(sheet_rows) < 2:
                continue

            # Find header row (the one containing "CLIENT")
            header_idx = None
            for i, row in enumerate(sheet_rows):
                if any(str(c).strip().upper() == 'CLIENT' for c in row if c is not None):
                    header_idx = i
                    break

            if header_idx is None:
                continue

            # Build column name mapping
            header = [str(c).strip().upper() if c is not None else f'_COL{j}'
                      for j, c in enumerate(sheet_rows[header_idx])]

            # Find indices of required columns
            col_map = {}
            for col_name in REQUIRED_COLS:
                if col_name in header:
                    col_map[col_name] = header.index(col_name)

            if 'CLIENT' not in col_map:
                continue

            ci = col_map['CLIENT']
            pi = col_map.get('PCS')
            wi = col_map.get('WASHING')
            wei = col_map.get('WASH ED')
            di = col_map.get('DATE')
            li = col_map.get('LOT NO.')

            # Process data rows
            for row in sheet_rows[header_idx + 1:]:
                ncols = len(row)

                client = row[ci] if ci < ncols else None
                if client is None:
                    continue
                client = str(client).strip()
                if not client or client == 'nan':
                    continue

                # filter by DATE column if present (keep from Nov 2025 inclusive)
                if di is not None and di < ncols and row[di] is not None:
                    try:
                        row_date = pd.to_datetime(row[di], errors='coerce')
                    except Exception:
                        row_date = pd.NaT
                    if pd.isna(row_date) or row_date < _date_cutoff:
                        continue

                pcs = 0
                if pi is not None and pi < ncols and row[pi] is not None:
                    try:
                        pcs = int(float(row[pi]))
                    except (ValueError, TypeError):
                        pcs = 0

                washing = ''
                if wi is not None and wi < ncols and row[wi] is not None:
                    washing = str(row[wi]).strip()
                    if washing == 'nan':
                        washing = ''

                wash_ed = ''
                if wei is not None and wei < ncols and row[wei] is not None:
                    wash_ed = str(row[wei]).strip()
                    if wash_ed == 'nan':
                        wash_ed = ''

                lot_no = ''
                if li is not None and li < ncols and row[li] is not None:
                    lot_no = str(row[li]).strip()
                    if lot_no == 'nan':
                        lot_no = ''

                washing_empty = washing == ''
                wash_ed_empty = wash_ed == ''

                making = pcs if washing_empty else 0
                in_washing = 0 if (washing_empty or not wash_ed_empty) else pcs
                out_washing = pcs if (not washing_empty and not wash_ed_empty) else 0

                all_rows.append({
                    'CLIENT': client,
                    'WASHING': washing,
                    'PCS': pcs,
                    'MAKING': making,
                    'IN_WASHING': in_washing,
                    'OUT_WASHING': out_washing,
                    'LOT_NO': lot_no,
                })
    finally:
        wb.close()

    return all_rows


def _read_sheet_fast(ws):
    """
    Read rows from a worksheet, stopping early once we detect
    the real data has ended (consecutive rows with only col 0 filled).
    """
    empty_streak = 0
    found_data = False

    for row in ws.iter_rows(values_only=True):
        # Count non-None values beyond column 0
        has_multi_col_data = any(c is not None for c in row[1:])

        if has_multi_col_data:
            found_data = True
            empty_streak = 0
            yield row
        else:
            empty_streak += 1
            if found_data and empty_streak > MAX_EMPTY_STREAK:
                break
            yield row


class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        start_time = time.time()
        try:
            excel_bytes = download_excel_bytes()
            load_time = time.time() - start_time

            process_start = time.time()
            all_rows = process_all_sheets(excel_bytes)
            process_time = time.time() - process_start

            if not all_rows:
                response_data = {
                    "rows": [],
                    "client_summary": [],
                    "washer_summary": [],
                    "total_pcs": 0,
                    "total_making": 0,
                    "total_in_washing": 0,
                    "total_out_washing": 0,
                    "timestamp": datetime.now().isoformat(),
                    "cached": _cache['timestamp'] is not None,
                    "processing_time": round(time.time() - start_time, 2),
                }
                self._send_json(response_data, 200)
                return

            # Build summaries using pandas (fast on small data)
            summary_start = time.time()
            master = pd.DataFrame(all_rows)

            # Client summary
            client_summary = (
                master.groupby('CLIENT', as_index=False, sort=False)
                .agg({'MAKING': 'sum', 'IN_WASHING': 'sum', 'OUT_WASHING': 'sum'})
                .sort_values('MAKING', ascending=False)
            )
            # Add TOTAL = MAKING + IN_WASHING + OUT_WASHING for client stats/graphs
            client_summary['TOTAL'] = client_summary[['MAKING', 'IN_WASHING', 'OUT_WASHING']].sum(axis=1)

            # Washer summary
            washer_data = master[master['WASHING'].str.len() > 0]
            if not washer_data.empty:
                washer_summary = (
                    washer_data.groupby('WASHING', as_index=False, sort=False)
                    .agg({'IN_WASHING': 'sum', 'OUT_WASHING': 'sum'})
                    .rename(columns={'WASHING': 'WASHER'})
                )
                # PENDING = items currently at washer (IN_WASHING represents items waiting to be completed)
                washer_summary['PENDING'] = washer_summary['IN_WASHING']
                # Add TOTAL = IN_WASHING + OUT_WASHING (total pieces handled by washer)
                washer_summary['TOTAL'] = washer_summary['IN_WASHING'] + washer_summary['OUT_WASHING']
                washer_summary = washer_summary.sort_values('PENDING', ascending=False)
            else:
                washer_summary = pd.DataFrame(columns=['WASHER', 'IN_WASHING', 'OUT_WASHING', 'PENDING', 'TOTAL'])

            # Breakdown: group by CLIENT + WASHING, sum PCS/MAKING/IN_WASHING/OUT_WASHING, collect LOT_NOs
            breakdown = (
                master.groupby(['CLIENT', 'WASHING'], as_index=False, sort=False)
                .agg({
                    'PCS': 'sum',
                    'MAKING': 'sum',
                    'IN_WASHING': 'sum',
                    'OUT_WASHING': 'sum',
                    'LOT_NO': lambda x: ', '.join(sorted(set(v for v in x if v)))
                })
            )
            # Add lot count
            breakdown['LOT_COUNT'] = breakdown['LOT_NO'].apply(lambda x: len(x.split(', ')) if x else 0)
            breakdown = breakdown.sort_values(['CLIENT', 'PCS'], ascending=[True, False])

            summary_time = time.time() - summary_start

            total_time = time.time() - start_time
            response_data = {
                "rows": breakdown.to_dict(orient='records'),
                "client_summary": client_summary.to_dict(orient='records'),
                "washer_summary": washer_summary.to_dict(orient='records'),
                "total_pcs": sum(r['PCS'] for r in all_rows),
                "total_making": sum(r['MAKING'] for r in all_rows),
                "total_in_washing": sum(r['IN_WASHING'] for r in all_rows),
                "total_out_washing": sum(r['OUT_WASHING'] for r in all_rows),
                "timestamp": datetime.now().isoformat(),
                "cached": (_cache['timestamp'] is not None and time.time() - _cache['timestamp'] < CACHE_DURATION),
                "processing_time": round(total_time, 2),
                "timing": {
                    "load": round(load_time, 2),
                    "process_sheets": round(process_time, 2),
                    "summaries": round(summary_time, 2),
                }
            }
            self._send_json(response_data, 200)

        except Exception as e:
            self._send_json({"error": str(e)}, 500)

    def _send_json(self, data, status):
        self.send_response(status)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(json.dumps(data).encode())